import pyautogui
import openpyxl
import os
import time

#submitButton = (100,100)
#submitButtonColor = (100,100,100)

# Anlegen der neuen 150er Rezeptur


#  im Teilestamm
#    Definitionen in Excel Tabelle hinterlegen
#    Aufruf über Terminal Sitzung, da TREND Client teilweise zu langsam ist
#    erstes SET muss vorher angelegt sein
#    Achtung Stammdaten 
#       Bezeichnung 1 muss nach neuer Systematik gefüllt werden
#       Bezeichnung 2 des zu kopierenden Sets leeren
#       SIBE sollte leer sein
#       Lager-Stamm Satz Sicherheitsbestand auf 0 setzen bzw. bei Neuanlage Satz löschen
#       Bruttogewicht prüfen
#       Referenz zum alten Set in EK-Gebinde eintragen 
#    Erste Maske muss voher manuell gestartet werden
#  Bei der SET Anlage wird 


# Excel Tabellenblatt laden



print('IdentNummer eingeben :') ;  IdentNummer = input()
print(IdentNummer)


# Excel Tabellenblatt laden
os.getcwd()
os.chdir('C:\\Programmcode\\Datenaustausch')
wb = openpyxl.load_workbook('Artikel.xlsx')
wb.get_sheet_names()
sheet = wb.get_sheet_by_name(IdentNummer)
sheet
type(sheet)
sheet.title
sheet.cell(row=1, column=1) # GebindeVariante
sheet.cell(row=1, column=2) # GebindeBezeichnung
sheet.cell(row=1, column=3) # VariantenBezeichnung
sheet.cell(row=1, column=4) # RAL
sheet.cell(row=1, column=5) # NCS
sheet.cell(row=1, column=6) # FarbName
sheet.cell(row=1, column=7) # SlvaVariante
sheet.cell(row=1, column=8).value # SetVariante



#  MCSKAK01
#positionieren des Mauszeigers in Maske
pyautogui.click(100, 100) 
# TODO Kopieren eines alten Artikels MCTEIL01 / dies muss vorher immer manuell erfolgen

pyautogui.typewrite('3', 0.25) #in feld klicken und kopieren eingeben

print('nach Aufruf erste Maske')

time.sleep(0.20)

pyautogui.typewrite('\n', 0.5) # Mit Enter weiter um nächste Maske aufzurufen

print('nach Aufruf zweite Maske')

#  Maske  ausfüllen TRTEILK1
# im range kann bei Programmabbuch oder Pause wieder neu aufgesetzt werden. Dieser WErt ist manuell zu ändern 
time.sleep(0.20)

print('vor ausfüllen MAske')

for i in range(1,2,1):
    SetVariante = (sheet.cell(row=i, column=8).value) 
    SlvaVariante = (sheet.cell(row=i, column=7).value) 
    VariantenBezeichnung = (sheet.cell(row=i, column=3).value)
    GebindeVariante = (sheet.cell(row=i, column=1).value)

    print('VariantenBezeichnung: ' + VariantenBezeichnung + '    SetVariante:    ' + SetVariante)
    
    # TRSKAKK1
    pyautogui.typewrite('150', 0.25)
    pyautogui.typewrite('\t')
    pyautogui.typewrite(str(SlvaVariante))
    pyautogui.typewrite('\t')
    pyautogui.typewrite('J')
    time.sleep(2.0)
    pyautogui.typewrite('\n', 0.5)

    # TRSKAK01
    pyautogui.typewrite('\t')
    pyautogui.typewrite(VariantenBezeichnung)
    pyautogui.typewrite('\n', 0.5)

    # MCSPAP01
    # TODO Anpassungen bei jeder Rezeptur prüfen
    pyautogui.typewrite('\t')
    pyautogui.typewrite('2', 0.25)
    pyautogui.typewrite('\n', 0.5)

    # TRSLP001
   # pyautogui.click(100,100)
   # pyautogui.doubleClick(100, 100) 
   # pyautogui.typewrite('\t')
    pyautogui.press('up')
    pyautogui.press('up')
    pyautogui.press('up')
    pyautogui.typewrite(GebindeVariante)
    print(GebindeVariante)
    print('TRSLP001 p113')
    time.sleep(2.0)
    pyautogui.press('f4')
    time.sleep(2.0)
    # MCTEIL01
    pyautogui.typewrite('1', 0.25)
    time.sleep(2.0)
    pyautogui.typewrite('\n', 0.5)
    print('TRSLP001 p113')

    # TRSLP001
    pyautogui.typewrite('\t')
    pyautogui.typewrite('\t')
    pyautogui.typewrite('\t')
    pyautogui.typewrite('01.10.2019')
    pyautogui.typewrite('\t')
    for j in range(1,11,1):    
        pyautogui.press('del')
    pyautogui.typewrite('\n', 0.5)
    print('TRSLP001 p113')
    # TRSLP002



    3
    pyautogui.typewrite('\n', 0.5)
    print('TRSLP001 p113')
    # MCSPAP01
    pyautogui.press('f12')

    # TRSLRK01
    pyautogui.typewrite('\n', 0.5)

    # TRSLRK62
    pyautogui.typewrite('\n', 0.5)

    # TRSLRK03
    pyautogui.typewrite('\n', 0.5)

    # MCSKAK01
    time.sleep(5)
    #  Umstellung auf Terminaleingabe durchführen
    pyautogui.typewrite('3', 0.25) #in feld 3 für kopieren eingeben
    pyautogui.typewrite('\n', 0.5) # Mit Enter weiter um nächste Maske aufzurufen


# ##########################################

        


# pyautogui.press('f7') # mit F7 speichern
# pyautogui.press('f12') # mit F12 Maske verlassen

# TODO Maske dynamisch über Liste und Zähler ausfüllen

print('ENDE')

