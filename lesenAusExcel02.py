import openpyxl
import os

os.getcwd()
os.chdir('C:\\Programmcode\\Datenaustausch')


wb = openpyxl.load_workbook('Import01.xlsx')
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Tabelle1')

sheet
type(sheet)

sheet.title
sheet.cell(row=1, column=1)
sheet.cell(row=1, column=2).value
for i in range(1,151,1):
    spalte1 = (sheet.cell(row=i, column=1).value) 
    spalte2 = (sheet.cell(row=i, column=2).value)
    print('spalte1: ' + spalte1 + '    spalte2:    ' + spalte2)


